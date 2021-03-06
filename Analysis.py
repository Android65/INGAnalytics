# -*- coding: utf-8 -*-
from collections import Counter
import os
import time
import json
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import NamedStyle, builtins, Alignment
from openpyxl.worksheet.filters import SortCondition
import openpyxl.utils as util
import csv



class Analysis:
    def __init__(self,personlist,conversationlist,utterancelist,startDate,usecasedict=None):
        self.personlist = personlist
        self.conversationlist = conversationlist
        self.utterancelist = utterancelist
        self.usecasedict = usecasedict
        self.wb = Workbook()
        self.conversationIDDict = {}
        self.startDate = startDate

    #This bundles all Utterance analysis methods, this method will also generate the JSON file for the frontend
    def analyseUtterances(self):
        self.countUsecases()
        self.countIntents()
        self.countWrong()

    def formatName(self,fName,lName):
        fName = str(fName) if fName is not None else ""
        lName = str(lName) if lName is not None else ""
        return fName+" "+lName

    def checkDate(self,uttTimeEpoch):
        pattern = '%Y%m%d'
        minEpoch = int(time.mktime(time.strptime(str(self.startDate), pattern)))
        #Returns True if the date is after the start date
        return True if minEpoch*1000<uttTimeEpoch else False

    def save(self):
        self.wb.save("Test.xlsx")

    def excelLayout(self,workbook,column_limit):
        wb1 = workbook
        # Set header style
        for col in wb1.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.style = builtins.styles["Accent1"]

        # Set Auto-Width
        for column_cells in wb1.columns:
            # Helper function to convert all data types to text
            def as_text(value):
                if value is None:
                    return ""
                return str(value)

            # Max text length of text in a certain column is determined to set the width
            length = max(len(as_text(cell.value)) for cell in column_cells)

            # If the length is too long, make it shorter and wrap text
            if length > column_limit:
                length = column_limit
                wb1.column_dimensions[column_cells[0].column].width = length
                for cell in column_cells:
                    cell.alignment = Alignment(wrapText=True)
            else:
                wb1.column_dimensions[column_cells[0].column].width = length

        filtersize = "A1:"+util.get_column_letter(wb1.max_column)+str(wb1.max_row)
        wb1.auto_filter.ref = filtersize

    def getConversationsExcel(self):
        wb1 = self.wb.active
        #Set headers
        wb1.title= "Conversations"
        Headers = ["Timestamp","User Name","Conversation ID","User Utterance","Detected Intent","Marie Reply",
                   "Confidence Score", "Sentiment Score", "Person ID"]
        wb1.append(Headers)
        templist = []
        #Format data and add to temporary list, for sorting later
        with open('conversations.csv', 'w') as csvfile:
            writer = csv.writer(csvfile, delimiter=';', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            for person in self.personlist:
                for conversation in person.convList:
                    for utterance in conversation.utteranceList:
                        if self.checkDate(utterance.timeStamp) is False:
                            continue
                        #Format Timestamp
                        newstamp = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(utterance.timeStamp/1000))
                        #Format name
                        name = self.formatName(person.fName,person.lName)
                        #Format replies
                        replies = ""
                        for x in utterance.replyUtt: replies += (x)
                        replies = replies.replace("\r","").replace("\n","")
                        #Format input utterance
                        inpututt = utterance.inputUtt.replace("\n","").replace("\r","")
                        #Write Line
                        writelist = [newstamp,name,conversation.id,inpututt,utterance.intentName,replies,
                                    utterance.intentScore,utterance.sentimentScore,person.id]
                        templist.append(writelist)
                        #writer.writerow(writelist)
                        


        #Sort entries by conversation ID
        templist.sort(key=lambda x: x[2])
        #Build Excel and list holding locations of each conversationID
        convID = -1
        for idx,entry in enumerate(templist):
            if entry[2] > convID:
                convID = entry[2]
                self.conversationIDDict[convID] = idx+2

            wb1.append(entry)
        self.excelLayout(wb1,85)

    def buildHyperlinks(self,workbook,idlocation):
        # Build hyperlinks
        for row in workbook.iter_rows(row_offset=1):
            if row[idlocation].value is None:
                break
            rowindex = self.conversationIDDict[int(row[idlocation].value)]
            link = "Test.xlsx#Conversations!C" + str(rowindex)
            row[idlocation].hyperlink = (link)

    def getFallbackConversationsExcel(self):
        wb1 = self.wb.create_sheet("Fallback Conversations")
        Headers = ["Conversation ID","Message Time","Previous Message","Previous Message Intent","Previous Message Reply","Last Message","Last Message Intent"]
        CONVERSATIONIDINDEX = 0
        wb1.append(Headers)
        templist = []
        with open('fallback.csv', 'w') as csvfile:
            writer = csv.writer(csvfile, delimiter=';', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            for conversation in self.conversationlist:
                conversation.utteranceList.sort(key=lambda x: x.timeStamp, reverse=False)
                #Checks for start date
                if self.checkDate(conversation.utteranceList[0].timeStamp) is False:
                    continue
                if conversation.utteranceList[0].intentName == "Default_Fallback_Intent":
                    timestamp = time.localtime(int(conversation.utteranceList[0].timeStamp / 1000))
                    prevStamp = time.strftime('%Y-%m-%d %H:%M:%S', timestamp)
                    templist.append([conversation.id, prevStamp, "None", "None", "None", conversation.utteranceList[0].inputUtt,conversation.utteranceList[0].intentName])
                for i in range(len(conversation.utteranceList) - 1):
                    prevUtt = conversation.utteranceList[i]
                    timestamp = time.localtime(int(prevUtt.timeStamp / 1000))
                    prevStamp = time.strftime('%Y-%m-%d %H:%M:%S', timestamp)
                    currUtt = conversation.utteranceList[i + 1]
                    if "Default_Fallback_Intent" in currUtt.intentName:
                        writelist = [conversation.id,prevStamp,prevUtt.inputUtt,prevUtt.intentName,str(prevUtt.replyUtt),currUtt.inputUtt,currUtt.intentName]
                        templist.append(writelist)
                        #writer.writerow(writelist)
        templist.sort(key = lambda x: x[CONVERSATIONIDINDEX])
        for element in templist:
            wb1.append(element)

        self.buildHyperlinks(wb1,CONVERSATIONIDINDEX)
        self.excelLayout(wb1,80)

    def getFailedConversationsExcel(self):
        wb1 = self.wb.create_sheet("Failed Dialog Flows")
        Headers = ["Usecase", "Conversation ID", "Time", "User Name", "Person ID", "Fail/Success", "Exit Intent"]
        CONVERSATIONIDINDEX = 1
        wb1.append(Headers)
        templist = []
        intentDict = self.categorizeIntents()
        with open('failed.csv', 'w') as csvfile:
            writer = csv.writer(csvfile, delimiter=';', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            for person in self.personlist:
                for conversation in person.convList:
                    conversation.utteranceList.sort(key=lambda x: x.timeStamp, reverse=False)
                    # Checks for start date
                    if self.checkDate(conversation.utteranceList[0].timeStamp) is False:
                        continue
                    intentList = []
                    for utterance in conversation.utteranceList:
                        intentList.append(utterance.intentName)
                    for usecase in intentDict:
                        entryintents = intentDict[usecase]["Intents"]["Start"]
                        exitintents = intentDict[usecase]["Intents"]["End"]
                        exitintents.extend(intentDict[usecase]["Intents"]["Jump"])
                        if not set(intentList).isdisjoint(entryintents):
                            newstamp = time.strftime('%Y-%m-%d %H:%M:%S',
                                                    time.localtime(conversation.firstUttTime / 1000))
                            name = self.formatName(person.fName, person.lName)
                            if not set(intentList).isdisjoint(exitintents):
                                writelist = [usecase,conversation.id, newstamp, name, person.id, "Succes",
                                            ", ".join(list(set(intentList) & set(exitintents)))]
                                templist.append(writelist)
                                #writer.writerow(writelist)
                            else:
                                writelist = [usecase, conversation.id, newstamp, name, person.id, "Fail",conversation.exitIntent]
                                templist.append(writelist)
                                #writer.writerow(writelist)

        #Sort list by conversation ID
        templist.sort(key=lambda x: x[CONVERSATIONIDINDEX])
        #Build excel
        for element in templist:
            wb1.append(element)

        #Build hyperlinks
        self.buildHyperlinks(wb1,CONVERSATIONIDINDEX)

        #All this is not very flexible, I know. Sorry future programmer! :(
        self.excelLayout(wb1, 80)

    def addToUseCaseDict(self, usecasedict, intentname, contextname, category):
        for contextkey, value in usecasedict.items():
            if value["ContextName"] == contextname:
                usecasedict[contextkey]["Intents"][category].append(intentname)
                return usecasedict

    def contextToUseCase(self,contextnamelist):
        usecaselist = []
        for item in contextnamelist:
            usecaselist.append(item[4:].replace("_", " ").capitalize())
        return usecaselist

    def buildActionsAndEvents(self, actionDict, eventDict, intentname, jsondict):
        # Build action and events dicts
        try:
            event = jsondict["events"][0]["name"]
            if event in eventDict:
                print("ERROR: EVENT OCCURS TWICE")
                print(eventDict[event])
                print(intentname)
                print("END ERROR")
            if event not in eventDict:
                eventDict[event] = intentname
        except:
            pass
        try:
            action = jsondict["responses"][0]["action"]
            if "fireevent:" in action:
                action = action[10:]
                # the [10:] is used to cut off the fireevent: part, leaving only the event that it connects to
                if action not in actionDict:
                    actionDict[action] = []
                actionDict[action].append(intentname)
        except Exception as e:
            pass

        return actionDict, eventDict

    # This method is meant to correct the fact that jump destinations are not logged. Therefore, if an intent jumps to
    # an exit intent, it should also be an exit intent
    def correctInternalExitJumps(self,usecasedict,actiondict,eventdict):
        #First we build a list of jumps. This list contains a ([From],To) tuple.
        jumpTupleList = []
        for eventname in set(actiondict).intersection(eventdict):
            jumpTupleList.append((actiondict[eventname],eventdict[eventname]))

        # Now we go over all the jumps and check for internal jumps
        for jumpTuple in jumpTupleList:
            for usecase in usecasedict:
                # If an intermediate intent jumps to an end intent, it becomes an end intent
                if jumpTuple[1] in usecasedict[usecase]["Intents"]["End"] and not set(jumpTuple[0]).isdisjoint(usecasedict[usecase]["Intents"]["Intermediate"]):
                    for element in jumpTuple[0]:
                        if element in usecasedict[usecase]["Intents"]["Intermediate"]:
                            usecasedict[usecase]["Intents"]["End"].append(element)
                            usecasedict[usecase]["Intents"]["Intermediate"].remove(element)
                # If a start intent jumps to an end intent, it also becomes an end intent
                if jumpTuple[1] in usecasedict[usecase]["Intents"]["End"] and not set(jumpTuple[0]).isdisjoint(usecasedict[usecase]["Intents"]["Start"]):
                    for element in jumpTuple[0]:
                        if element in usecasedict[usecase]["Intents"]["Start"]:
                            usecasedict[usecase]["Intents"]["End"].append(element)

        return usecasedict

    def mergeOrderCardIntent(self,usecasedict):
        rydict = usecasedict["Unblock card ry"]["Intents"]
        rndict = usecasedict["Unblock card rn"]["Intents"]
        #This merges unblock card ry and rn into unblock card
        usecasedict["Unblock card"]["Intents"] = {key: value + rydict[key] + rndict[key] for key, value in usecasedict["Unblock card"]["Intents"].items()}
    
        return usecasedict

    def categorizeIntents(self):
        generalUseCaseContextList = ["ctx_activate_card", "ctx_sms_authentication", "ctx_card_delivery",
                                     "ctx_card_stop",
                                     "ctx_info_charges_abroad", "ctx_newpin", "ctx_order_card", "ctx_pin_by_sms",
                                     "ctx_unblock_card", "ctx_unblock_card_ry", "ctx_unblock_card_rn", "ctx_use_abroad"]
        generalUseCaseContextSet = set(generalUseCaseContextList)
        usecaseNameList = self.contextToUseCase(generalUseCaseContextList)

        # Build the dictionary that will be converted to JSON in the end
        finaldict = {}
        for usecase, context in zip(usecaseNameList, generalUseCaseContextList):
            finaldict[usecase] = {"ContextName": context, "Intents":
                {"Start": [], "Intermediate": [], "End": [], "Fallback": [], "Jump": []}}

        # Additionally, we define two dicts: one holding all intents with events, and the other with actions. These are used to create the jump list
        actionDict = {}
        eventDict = {}

        dir = os.getcwd()
        intentdir = os.path.join(dir,"Intents")

        for filename in os.listdir(intentdir):
            intentname = filename.split(".")[0]

            # Open the file
            with open(os.path.join(intentdir,filename), "r", encoding='utf8') as file:
                hasinput = True
                hasoutput = True

                jsondict = json.load(file)

                #If a certain intent has an action or intent, add it to the relevant dicts
                actionDict,eventDict = self.buildActionsAndEvents(actionDict, eventDict, intentname, jsondict)

                # Define input and output contexts
                outcontextsdict = jsondict["responses"][0]["affectedContexts"]
                outcontextslist = []
                for context in outcontextsdict:
                    if context['lifespan'] > 0:
                        outcontextslist.append(context["name"])
                incontextslist = jsondict["contexts"]

                # See if any of the general contexts are present in either the input or output contexts
                # If it is present, save the fallback
                if generalUseCaseContextSet.isdisjoint(incontextslist):
                    hasinput = False
                else:
                    incontext = next(iter(generalUseCaseContextSet.intersection(incontextslist)))

                if generalUseCaseContextSet.isdisjoint(outcontextslist):
                    hasoutput = False
                else:
                    outcontext = next(iter(generalUseCaseContextSet.intersection(outcontextslist)))

                '''
                We use the following truth table to determine what the intent is:
                Input/Output    Intent Category         Remarks

                False/False     Generic intent
                False/True      Start intent
                True/False      End intent              Checks end_of_conversation to be sure
                True/True       Intermediate intent*    *if the context is not the same in input and output, it's an exit
                                                        because it jumps from one use case to another
                '''

                if (not hasinput) and (not hasoutput):
                    continue
                elif (not hasinput) and hasoutput:
                    finaldict = self.addToUseCaseDict(usecasedict=finaldict, intentname=intentname,
                                                      contextname=outcontext, category="Start")
                elif hasinput and (not hasoutput):
                    if "fallback" in intentname.lower() and hasinput:
                        finaldict = self.addToUseCaseDict(usecasedict=finaldict, intentname=intentname,
                                                     contextname=incontext, category="Fallback")
                        continue
                    if "ctx_end_of_conversation" in outcontextslist:
                        finaldict = self.addToUseCaseDict(usecasedict=finaldict, intentname=intentname,
                                                     contextname=incontext, category="End")
                    else:
                        print("EXIT INTENT WITHOUT END OF CONVERSATION: " + intentname)
                elif hasinput and hasoutput:
                    if incontext == outcontext:
                        finaldict = self.addToUseCaseDict(usecasedict=finaldict, intentname=intentname,
                                                     contextname=incontext, category="Intermediate")
                    else:
                        finaldict = self.addToUseCaseDict(usecasedict=finaldict, intentname=intentname,
                                                     contextname=incontext, category="Jump")

        # Here we correct the internal jumps to exit intents
        finaldict = self.correctInternalExitJumps(finaldict,actionDict,eventDict)

        # Finally, we merge the order card ry and rn into order card
        finaldict = self.mergeOrderCardIntent(finaldict)

        return finaldict

    def getPersons(self):
        for person in self.personlist:
            print(person.fName)
            print(person.lName)
            print("-------------------")


'''
    def cleanNetworkEdges(self,intentlist):
        for index, item in enumerate(intentlist):
            if "Generic" in item:
                intentlist[index] = "Generic"
        return intentlist


    def buildNetworkEdges(self, G):
        vertexList = []
        for conversation in self.conversationlist:
            intentlist = self.cleanNetworkEdges(conversation.intentList)
            for i in range(len(intentlist)-1):
                vertexList.append((intentlist[i],intentlist[i+1]))
        vertexcount = dict(Counter(vertexList))
        for vertex,count in vertexcount.items():
            try:
                G.add_edge(*vertex, weight=count)
            except:
                print("fuck")
        return G

    def buildNodes(self):
        G = nx.DiGraph()
        yStart = 1
        xPos = 1
        yMax = 0
        for usecase,intentlist in self.usecasedict.items():
            yPos = yMax
            xPos = 1
            print("next")
            posDict = {}
            for intent in intentlist:
                try:
                    if(len(intentlist)>1):
                        s = int(''.join(x for x in intent if x.isdigit()))
                    else:
                        s = 0
                except ValueError:
                    s = 0
                if s > xPos-1:
                    xPos += 1
                    yPos = yStart
                elif s < xPos-1:
                    xPos = s +1
                    yPos = yStart+1
                else:
                    yPos += 1
                if yPos > yMax:
                        yMax = yPos
                print((xPos,yPos))
                posDict[intent]=(xPos,yPos)
                G.add_node(intent,pos=(xPos,yPos))
            yStart = yMax + 2
            yMax += 1
        return G,posDict

    def visualizeNetwork(self,G,pos):
        edge_trace = Scatter(
            x=[],
            y=[],
            line=Line(width=0.5, color='#888'),
            hoverinfo='text',
            text= [],
            mode='lines')

        for edge in G.edges():
            print(edge)
            x0, y0 = pos[edge[0]]
            x1, y1 = pos[edge[1]]
            edge_trace['text'].append(edge["weight"])
            edge_trace['x'] += [x0, x1, None]
            edge_trace['y'] += [y0, y1, None]

        node_trace = Scatter(
            x=[],
            y=[],
            text=[],
            textposition="center right",
            mode='markers+text',
            hoverinfo='all',
            marker=Marker(
                showscale=True,
                # colorscale options
                # 'Greys' | 'Greens' | 'Bluered' | 'Hot' | 'Picnic' | 'Portland' |
                # Jet' | 'RdBu' | 'Blackbody' | 'Earth' | 'Electric' | 'YIOrRd' | 'YIGnBu'
                colorscale='YIGnBu',
                reversescale=True,
                color=[],
                size=10,
                line=dict(width=2)))
        for node in G.nodes():
            try:
                x, y = pos[node]
                node_trace['text'].append(str(node))
                node_trace['x'].append(x)
                node_trace['y'].append(y)
            except Exception as e:
                print(e)
                print(node)


        fig = Figure(data=Data([edge_trace,node_trace]),
                     layout=Layout(
                         title='<br>Network graph made with Python',
                         titlefont=dict(size=16),
                         showlegend=False,
                         hovermode='closest',
                         margin=dict(b=20, l=5, r=5, t=40),
                         xaxis=XAxis(showgrid=True, zeroline=True, showticklabels=False),
                         yaxis=YAxis(showgrid=True, zeroline=True, showticklabels=False)))
        po.plot(fig, filename='networkx.html')

    def buildGraph(self):
        G,initPosition = self.buildNodes()
        G = self.buildNetworkEdges(G)
        pos = nx.spring_layout(G,pos=initPosition)
        self.visualizeNetwork(G,pos)

'''
